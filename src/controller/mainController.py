from  src.model.mainModel import MainModel
from src.view.mainView import MainView
from styles import SetupTheme
from openpyxl import Workbook
from openpyxl import load_workbook
from src.android.androidDeviceManager import AndroidDeviceManager
from PySide6.QtCore import QThread, Signal
import re
from src.controller.modalCreateAutoController import ModalCreateAutoController



class AdbThread(QThread):

    finished = Signal(str)
    error = Signal(str)
    progress = Signal(int)


    def __init__(self, adb_manager:AndroidDeviceManager, phones_number:list, message:list, parent=None):
        super().__init__(parent)
        self.adb_manager = adb_manager
        self.phones_number:list = phones_number
        self.messages:list = message
        self.stop = False

    def run(self):
        try:
            #--------------------------------
            self.stop = False
            self.adb_manager.screen_time_on_5min()
            self.adb_manager.habiliar_adbkeyboard()

            for index, phone in enumerate(self.phones_number, start=1):

                progress_int = int((index / len(self.phones_number)) * 100)
                self.progress.emit(progress_int)

                if self.stop:
                    self.error.emit(f"STOP realizado com sucesso!")
                    break

                if self.is_valid_phone_number(phone):

                    for msg in self.messages:
                        status, mensagem = self.adb_manager.mensagem_whats(phone, msg)
                        if status:
                            self.finished.emit(f"[{phone}] : MENSAGEM ENVIADA COM SUCESSO 😊")
                        else:
                            self.error.emit(mensagem)
                            break
                else:
                    self.finished.emit(f"Telefone [VAZIO ou INVALIDO].Por favor, verificar!")
                
        except Exception as e:
            self.error.emit(str(e))

    def stop_script(self):
        self.stop = True

    def is_valid_phone_number(self,phone: str) -> bool:
        pattern = r'^(\(\d{2}\)\d{5}-\d{4}|\d{11}|\d{7}-\d{4})$'
        return bool(re.match(pattern, phone))




class MainController:
    
    def __init__(self):
        #----------------------------
        self.main_model = MainModel()
        self.main_view = MainView()
        self.theme = SetupTheme()
        self.theme.setupTheme('dark')
        self.theme_select()
        self.setup_connections()
        self.setup_connections_main_window()
        self.main_view.show()

        #------------------------ variaveis de controller -----------------
        self.value_send_phone = None
        self.data_plan = None
        self.adb = None
        self.modal = None
        self.modal_create_controller:ModalCreateAutoController = None


    def setup_connections_main_window(self):
        self.main_view.open_action.triggered.connect(self.open_action_file) 
        self.main_view.combo_box.currentTextChanged.connect(self.theme.setupTheme)
        self.main_view.tree_widget.itemClicked.connect(self.on_item_clicked_tree)

    def setup_connections(self):
        self.main_view.start_process.clicked.connect(self.start_process)
        self.main_view.stop_process.clicked.connect(self.stop_process)
        self.main_view.combo_box_colun_envio_phone.currentTextChanged.connect(self.text_selection_phone)



    def on_item_clicked_tree(self, item, column):

        if item.text(column) == "Whatsapp":
            self.main_view.central_widget_main()
            self.setup_connections()
           
        elif item.text(column) == "Modal-de-criacao":
            modal_create_auto = self.modal_create_controller.get_modal_create_auto_widget()
            self.main_view.setCentralWidget(modal_create_auto)
            

    def open_action_file(self):
        file_path_xlsx = self.main_view.open_action_file()

        if file_path_xlsx is not None:
            self.data_plan = self.get_planilha(file_path_xlsx)
            self.main_view.add_value_combo_box_envio_phone(self.data_plan.get('header_label'))
            self.main_view.show_table_widget_view(data=self.data_plan)

        
    def start_process(self):
        telefones = self.main_view.get_column_data(self.value_send_phone)

        if telefones:
            text = self.main_view.message_text.toPlainText()
            
            if text.strip() == '':
                self.log("MENSAGEM não encontrada!")
                return
            
            msg = []
            msg.append(text)
            
            #----------------------------------
            self.adb = AndroidDeviceManager()
            self.adb.connect_device()

            self.adb_thread = AdbThread(self.adb, telefones, msg)
            self.adb_thread.finished.connect(self.on_adb_finished)
            self.adb_thread.error.connect(self.on_adb_error)
            self.adb_thread.progress.connect(self.progress_send_message)
            self.adb_thread.start()
        else:
            self.log("NUMERO não encontrado!")

    def progress_send_message(self,value:int):
        self.main_view.progress_bar.setValue(value)

    def on_adb_finished(self,msg):
       self.log(msg=msg)

    def on_adb_error(self, error_message):
       self.log(f"{error_message}")
     


    def stop_process(self):
        if self.adb is not None:
           self.adb_thread.stop_script()

        
    def get_planilha(self,path):
        try:
            workbook = load_workbook(path)
            sheet = workbook.active
            
            header_label = []
            data = []

            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0: 
                    header_label = [cell if cell is not None else '' for cell in row]
                else: 
                    data.append([cell if cell is not None else '' for cell in row])
            
            result = {
                "header_label": header_label,
                "data": data
            }
            
            return result
        
        except Exception as e:
            print(f"Erro ao processar o arquivo: {e}")
            return None
        

    def create_planilha(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Minha Planilha"

        ws.append(["NOME", "EMAIL", "TELEFONE","MENSAGEM"])
        ws.append([f"Francisco santos", "fra@gmail.com", "92993160919","OLA AQUI È UM TESTE DE MENSAGEM QUE SERÀ ENVIADO"])

        # Salvar a planilha
        wb.save("exemplo.xlsx")
    

    def log(self,msg):
        self.main_view.log_view(msg=msg)

    def set_controller(self,controller):
        self.modal_create_controller = controller

    def theme_select(self):
        self.main_view.combo_box.addItems(self.theme.getTheme())

    def text_selection_phone(self,txt):
        self.value_send_phone = txt









    

        
